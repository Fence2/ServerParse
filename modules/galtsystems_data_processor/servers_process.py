from openpyxl import Workbook
from openpyxl.styles import Font

from modules.parser.tools import get_today_full_date_str, format_excel_columns_width, search_from_pattern, Patterns, \
    sub_not_digits


def process_servers(
        servers: list,

        other_servers: list,

        folder_to_save,
        filename: str
):
    import modules.galtsystems_data_processor.process_by_category as process_by_category

    servers = list(sorted(servers, key=lambda x: (x.category, x.config_price)))
    processed_data = process_by_category.servers.servers_process(servers)
    gs_servers, good_cfg, part_cfg, bad_cfg = processed_data

    wb = Workbook()
    ws_gs = wb.active
    ws_gs.title = "ГалтСистемс"  # noqa
    ws_gs.append(["Категория", "Название", "Цена"])

    ws_good = wb.create_sheet(title="Идеальный конфиг")

    # Первый лист: Сводная инфо по хорошим серверам, нужным нашей компании
    for k, v in gs_servers.items():
        ws_gs.append(["Сервер", k, v])
    format_excel_columns_width(ws_gs)

    # Остальные листы

    def write_cfg_to_book(cfg_servers: dict, sheet):
        rows = 0
        for server in cfg_servers.values():
            url = ''
            for i, line in enumerate(server):
                if not i:
                    url = line[0][1]
                    line[0] = line[0][0]

                sheet.append(line)

            if url != '':
                sheet.cell(row=rows + 1, column=1).hyperlink = url
                sheet.cell(row=rows + 1, column=1).style = "Hyperlink"
                sheet.cell(row=rows + 1, column=3).font = Font(bold=True)
                sheet.cell(row=rows + 1, column=4).font = Font(bold=True)

            rows += len(server)

    if len(good_cfg):
        write_cfg_to_book(good_cfg, ws_good)
        format_excel_columns_width(ws_good)
    if len(part_cfg):
        ws_part = wb.create_sheet(title="Не полный конфиг")
        write_cfg_to_book(part_cfg, ws_part)
        format_excel_columns_width(ws_part)
    if len(bad_cfg):
        ws_bad = wb.create_sheet(title="Плохой конфиг")
        write_cfg_to_book(bad_cfg, ws_bad)
        format_excel_columns_width(ws_bad)

    ws_all_servers = wb.create_sheet(title="Все серверы")
    ws_all_servers.append([
        'Title',
        'Brand',
        'Model',
        'Generation',
        'Units',
        'Trays amount',
        'Trays form-factor',
        'New',
        'Price'
    ])
    for server in servers + other_servers:
        trays_amount = search_from_pattern(Patterns.SERVER.trays, server.name)
        trays_form_factor = search_from_pattern(Patterns.FORM_FACTOR, server.name)
        ws_all_servers.append([
            server.name,
            server.brand,
            server.model,
            server.generation,
            server.units,
            sub_not_digits(trays_amount) if trays_amount is not None else '',
            trays_form_factor if trays_form_factor is not None else '',
            server.new,
            server.card_price
        ])

    format_excel_columns_width(ws_all_servers)

    filename = filename + "_" + get_today_full_date_str() + ".xlsx"

    wb.save(folder_to_save / filename)

    print("Успешное сохранение данных с парсера в файл: ", folder_to_save / filename)  # noqa
    return True
